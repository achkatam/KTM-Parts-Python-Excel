import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

headers = ["Part name", "Part number", "Price"]

for col, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col).value = header

parts_list = [
    ("Radiator fan", "79735941044", 131.99),
    ("Brake caliper support with disc guard", "79613975044", 173.99),
    ("Factory seat", "79707240000", 243.69),
    ("Front rotor guard", "7900996110030", 66.99),
    ("Supersprox sprocket", "5.84101E+12", 94.99),
    ("Support straps", "78712916000", 40),
    ("Chain protector", "78104974100", 99.99),
    ("Wrap around hand guard plastic", "79602979044", 35),
    ("Map switch", "51539974200", 94.99),
    ("Skid plate", "55403090144", 94.99),
    ("Clutch slave cylinder protection", "55432975044", 57.99),
    ("Solid rear brake rotor", "25010960000", 170.99)
]

for row, part in enumerate(parts_list, start=2):
    part_name, part_number, price = part
    sheet.cell(row=row, column=1).value = part_name
    sheet.cell(row=row, column=2).value = part_number
    sheet.cell(row=row, column=3).value = price

total_price = sum(price for _, _, price in parts_list)
sheet.cell(row=row + 1, column=2).value = "Total Price"
sheet.cell(row=row + 1, column=3).value = total_price

wb.save("part_list.xlsx")
